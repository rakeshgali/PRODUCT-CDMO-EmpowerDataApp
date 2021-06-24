# Version 12: Allow user input of detection limit (DL)

print('Running...\n')







# QL = 0.00  # For testing only
while True:
    QL = input('QL: ')
    if float(QL) > 0.2:
        print('QL > 0.2 is not allowed.\n')
    elif float(QL) < 0:
        print('QL cannot be negative\n')
    else:
        break
print(f'The QL you entered is {QL}\n')

import os

# get txt files
stop_program = False
file_list = []
for f_name in os.listdir('data'):
    if f_name.endswith('.txt'):
        file_list.append(f_name)
print(f'The list of txt file in this directory: {file_list}\n')
if len(file_list) == 0 or file_list[0] == 'error_message.txt':  # ensure txt file exported from Empower exist
    file = open('error_message.txt', 'w+')
    file.write(
        'No txt file exported from Empower was found.\nPlease remove this error_message txt file after reading this message.')
    file.close()
    stop_program = True

for file_name in range(len(file_list)):
    if file_list[file_name] == 'error_message.txt' and len(file_list) != 1:
        os.remove('error_message.txt')
        file = open('error_message.txt', 'w+')
        file.write('Please remove this error_message txt file before proceeding.')
        file.close()
        stop_program = True
        # print(f'File name "error_message.txt" was found')

import csv

RRT_value = 0
name_list = []
RRT_list = []
sample_name_list = []
Name_RRT_list = []
file_sample_list = []
## Create a master list of RRT with peak name and AN associated ##
if stop_program == False:
    for file in file_list:
        line_number = 0  # First line of txt file.
        with open(os.path.join('data', file)) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter='\t')
            for row in csv_reader:
                if line_number == 1:
                    file_name = file
                    sample_name = row[0]
                    file_sample_tuple = (file_name, sample_name)
                    file_sample_list.append(file_sample_tuple)
                elif line_number > 2:
                    AN = row[2]
                    if AN != '':
                        if float(AN) < float(QL):
                            AN = ''
                    if 'RRT~' in row[0] and AN != '':
                        Name = 'RRT~ ' + row[1]
                        RRT = row[1]
                        Name_RRT_tuple = (Name, RRT)
                        Name_RRT_list.append(Name_RRT_tuple)
                        if Name_RRT_tuple not in Name_RRT_list:
                            Name_RRT_list.append(Name_RRT_tuple)
                    elif AN != '' and row[0] != '' and row[1] != '':
                        Name = row[0]
                        RRT = row[1]
                        Name_RRT_tuple = (Name, RRT)
                        if Name_RRT_tuple not in Name_RRT_list:
                            Name_RRT_list.append(Name_RRT_tuple)
                line_number += 1

    file_sample_list_final = sorted(set(file_sample_list), key=lambda x: x[
        1])  # Sort by sample name and remove duplicates when both sample name and file name are the same
    Name_RRT_list_final = sorted(set(Name_RRT_list), key=lambda x: x[
        1])  # Sort by RRT and remove duplicates when both peak name and RRT are the same
    # print(Name_RRT_list_final)
    # print('\n')
    # from operator import itemgetter
    # Name_RRT_list_final = sorted(set(Name_RRT_list), key=itemgetter(1,0)) # Sort by RRT and remove duplicates when both peak name and RRT are the same
    # print(Name_RRT_list_final)
    # sample_name_list_final = sorted(sample_name_list)

    # print(f'Final Name_RRT_list is {Name_RRT_list}.\nThe number of entry is {len(Name_RRT_list)}')
    # print(f'Sorted Name_RRT_list is {Name_RRT_list_final}.\nThe number of entry is {len(Name_RRT_list_final)}')
    # print(f'Sorted file_sample_list is {file_sample_list_final}.\nThe number of entry is {len(file_sample_list_final)}')

    # Save sorted data into a new sheet and create a new file
    from openpyxl import Workbook
    from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, numbers
    from openpyxl.comments import Comment
    from openpyxl.utils import units

    wb = Workbook()
    new_ws = wb.active

    center_aligned_text = Alignment(horizontal="center")
    right_aligned_text = Alignment(horizontal="right")

    new_ws.title = 'Compiled Data'
    new_ws['B1'] = 'RT'
    new_ws.cell(row=1, column=2).font = Font(bold=True)
    new_ws.cell(row=1, column=2).alignment = right_aligned_text
    new_ws['B2'] = 'Lambda Max'
    new_ws.cell(row=2, column=2).font = Font(bold=True)
    new_ws.cell(row=2, column=2).alignment = right_aligned_text
    new_ws['B3'] = 'RRT'
    new_ws.cell(row=3, column=2).font = Font(bold=True)
    new_ws.cell(row=3, column=2).alignment = right_aligned_text
    new_ws['B4'] = 'Peak Name'
    new_ws.cell(row=4, column=2).font = Font(bold=True)
    new_ws.cell(row=4, column=2).alignment = right_aligned_text
    new_ws['A4'] = 'File Name'
    new_ws.cell(row=4, column=1).font = Font(bold=True)
    # new_ws.merge_cells('A1:A2')

    ### This section print the RRT and Peak Name into rows 3 and 4 ###
    ### Codes in v8 ###
    # x = 3  # second column
    # peak_number = len(Name_RRT_list_final)
    # print(f'Total number of unique peak in compiled data table is: {peak_number}')
    # for peak_index in range(peak_number):
    #     Name_RRT_tuple_ws = Name_RRT_list_final[peak_index]
    #     Name_ws = Name_RRT_tuple_ws[0]
    #     RRT_ws = Name_RRT_tuple_ws[1]
    #     new_ws.cell(row=3, column=x).value = float(RRT_ws)
    #     new_ws.cell(row=3, column=x).font = Font(bold=True)
    #     new_ws.cell(row=4, column=x).value = Name_ws
    #     new_ws.cell(row=4, column=x).font = Font(bold=True)
    #     new_ws.cell(row=3, column=x).alignment = center_aligned_text
    #     new_ws.cell(row=4, column=x).alignment = center_aligned_text
    #     if new_ws.cell(row=4, column=x).value == new_ws.cell(row=4, column=x-1).value:  # Highlight peak with same name but diff RRT; help user tidy up data
    #         # new_ws.cell(row=3, column=x).font = Font(color="00FF6600")
    #         new_ws.cell(row=4, column=x).font = Font(bold=True, color="00FF6600")
    #         # new_ws.cell(row=3, column=x-1).font = Font(color="00FF6600")
    #         new_ws.cell(row=4, column=x-1).font = Font(bold=True, color="00FF6600")
    #     if str(new_ws.cell(row=3, column=x).value) == str(new_ws.cell(row=3, column=x-1).value):  # Highlight peak with same RRT but diff name; help user tidy up data
    #         # new_ws.cell(row=2, column=x).font = Font(color="00FF6600")
    #         new_ws.cell(row=3, column=x).font = Font(bold=True, color="00FF6600")
    #         # new_ws.cell(row=2, column=x-1).font = Font(color="00FF6600")
    #         new_ws.cell(row=3, column=x-1).font = Font(bold=True, color="00FF6600")
    #     x = x + 1
    ### Codes in v8 ###
    ### End of section ###

    ## Create RRT and Name lists for Excel without duplication of Name ##
    x = 3  # third column
    excel_name_list = []
    excel_RRT_list = []
    peak_number = len(Name_RRT_list_final)
    for peak_index in range(peak_number):
        Name_RRT_tuple = Name_RRT_list_final[peak_index]
        Name = Name_RRT_tuple[0]
        RRT = Name_RRT_tuple[1]
        if peak_index == 0:
            excel_name_list.append(Name)
            excel_RRT_list.append(RRT)
            new_ws.cell(row=3, column=x).value = float(RRT)
            new_ws.cell(row=3, column=x).font = Font(bold=True)
            new_ws.cell(row=3, column=x).alignment = center_aligned_text
            new_ws.cell(row=4, column=x).value = Name
            new_ws.cell(row=4, column=x).font = Font(bold=True)
            new_ws.cell(row=4, column=x).alignment = center_aligned_text
            x = x + 1
        elif peak_index > 0:
            # previous_Name_RRT_tuple = Name_RRT_list_final[peak_index-1]
            # previous_Name = previous_Name_RRT_tuple[0]
            # previous_RRT = previous_Name_RRT_tuple[1]
            if Name in excel_name_list:
                name_index = excel_name_list.index(Name)
                RRT_index = name_index
                if new_ws.cell(row=3, column=3 + RRT_index).value == None:
                    RRT = excel_RRT_list[RRT_index]
                    new_ws.cell(row=3, column=3 + RRT_index).value = RRT
                    new_ws.cell(row=3, column=3 + RRT_index).number_format = numbers.FORMAT_NUMBER_00
                else:
                    existing_RRT = new_ws.cell(row=3, column=3 + RRT_index).value
                    RRT = str(existing_RRT) + ', ' + str(RRT)
                    new_ws.cell(row=3, column=3 + RRT_index).value = RRT
                new_ws.cell(row=3, column=3 + RRT_index).value = RRT
                new_ws.cell(row=3, column=3 + RRT_index).font = Font(bold=True)
                new_ws.cell(row=3, column=3 + RRT_index).alignment = center_aligned_text
            else:
                excel_name_list.append(Name)
                excel_RRT_list.append(RRT)
                new_ws.cell(row=3, column=x).value = float(RRT)
                new_ws.cell(row=3, column=x).font = Font(bold=True)
                new_ws.cell(row=3, column=x).alignment = center_aligned_text
                new_ws.cell(row=4, column=x).value = Name
                new_ws.cell(row=4, column=x).font = Font(bold=True)
                new_ws.cell(row=4, column=x).alignment = center_aligned_text
                x = x + 1

    print(f'excel_name_list is:{excel_name_list}\n')
    print(f'excel_RRT_list is:{excel_RRT_list}\n')
    ## End of this subsection ##

    ### This section print file name and sample name to column B ###
    y = 5  # starting row for file name and sample name population in Excel
    x = 3  # starting column for AN population in Excel
    file_count = len(file_sample_list_final)
    peak_number = len(excel_name_list)  # v9 code
    Do_Not_Overwrite_list = []
    for count in range(file_count):
        file_name_tuple = file_sample_list_final[count]
        file_name = file_name_tuple[0]
        sample_name = file_name_tuple[1]
        new_ws.cell(row=y, column=1).value = file_name
        # new_ws.cell(row=y, column=1).font = Font(bold=True)
        new_ws.cell(row=y, column=2).value = sample_name
        # new_ws.cell(row=y, column=2).font = Font(bold=True)
        ### End of section ###

        print(file_name)

        with open(os.path.join('data', file_name)) as csv_file:
        #with open(file_name) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter='\t')
            line_number = 0  # First line of txt file.
            for row in csv_reader:
                if line_number == 2:  # Count number of column in txt file
                    column_count = len(row)
                    data_type = row
                    # if 'Peak Lambda Max.' in data_type:0.2
                    #     lambda_max_exist = True
                    # else:
                    #     lambda_max_exist = False

                    # if 'Retention Time' in data_type:
                    #     RT_exist = True
                    # else:
                    #     RT_exist = False

                if line_number > 2:
                    Name = row[0]
                    RRT = row[1]
                    AN = row[2]
                    if AN != '':
                        if float(AN) < float(QL):
                            AN = ''
                            if float(
                                    RRT) == 1:  # handles special condition where after QL is applied, AN value for RRT = 1 becomes ND
                                Do_Not_Overwrite_list.append(file_name)
                    if 'RRT~' in row[0]:
                        Name = 'RRT~ ' + row[1]
                    x = 3

                    ##### v9 code #####
                    for peak_index in range(peak_number):
                        l_max_exist = True
                        RT_exist = True
                        Name_ws = excel_name_list[peak_index]  # for comparison purpose
                        ##### v9 code #####

                        ##### v8 code #####
                        # for peak_index in range(peak_number):
                        #    Name_RRT_tuple = Name_RRT_list_final[peak_index]
                        #    Name_ws = Name_RRT_tuple[0]  # for comparison purpose
                        #    RRT_ws = Name_RRT_tuple[1]  # for comparison purpose
                        ##### v8 code #####
                        # if lambda_max_exist == True:
                        existing_l_max_value = new_ws.cell(row=2, column=x).value
                        l_max = row[3]
                        if l_max == '':
                            l_max_exist = False
                        # if RT_exist == True:
                        existing_RT_value = new_ws.cell(row=1, column=x).value
                        RT = row[4]
                        if RT == '':
                            RT_exist = False

                        if Name == Name_ws and AN != '' and RRT != '':  # if Peak Name from compiled table matches that from each sample
                            if new_ws.cell(row=y,
                                           column=x).value != None:  # Deal with more than one value occupying same cell and highlight in red font
                                # print(f'Warning; cell already has a value, which is {new_ws.cell(row=y, column=x).value}')
                                existing_value = new_ws.cell(row=y, column=x).value
                                overlap_value = float(AN)
                                new_ws.cell(row=y, column=x).value = str(existing_value) + ', ' + str(overlap_value)
                                new_ws.cell(row=y, column=x).font = Font(color="00FF6600")

                                if l_max_exist == True:
                                    new_ws.cell(row=2, column=x).value = str(existing_l_max_value) + ', ' + str(l_max)
                                    new_ws.cell(row=2, column=x).alignment = center_aligned_text
                                    # new_ws.cell(row=1, column=x).font = Font(bold=True)
                                if RT_exist == True:
                                    new_ws.cell(row=1, column=x).value = str(existing_RT_value) + ', ' + str(RT)
                                    new_ws.cell(row=1, column=x).alignment = center_aligned_text
                            else:
                                new_ws.cell(row=y, column=x).value = float(
                                    AN)  # overwrite ND with AN value in number format if it exists
                                new_ws.cell(row=y, column=x).alignment = center_aligned_text
                                new_ws.cell(row=y, column=x).number_format = numbers.FORMAT_NUMBER_00

                                if l_max_exist == True and existing_l_max_value == None:
                                    # print(f'What is l_max at x={x} and y={y}: {l_max}')
                                    new_ws.cell(row=2, column=x).value = float(
                                        l_max)  # fill in first row with lamda max information
                                    # new_ws.cell(row=2, column=x).value = "{:.1f}".format(float(l_max))  # fill in first row with lamda max information
                                    new_ws.cell(row=2, column=x).alignment = center_aligned_text
                                    new_ws.cell(row=2, column=x).number_format = "0.0"
                                    # new_ws.cell(row=1, column=x).font = Font(bold=True)
                                elif l_max_exist == True and str(l_max) not in str(existing_l_max_value):
                                    new_ws.cell(row=2, column=x).value = str(existing_l_max_value) + ', ' + str(l_max)
                                    new_ws.cell(row=2, column=x).alignment = center_aligned_text
                                    # new_ws.cell(row=1, column=x).font = Font(bold=True)

                                if RT_exist == True and existing_RT_value == None:
                                    new_ws.cell(row=1, column=x).value = float(RT)
                                    new_ws.cell(row=1, column=x).alignment = center_aligned_text
                                    new_ws.cell(row=1, column=x).number_format = numbers.FORMAT_NUMBER_00
                                elif RT_exist == True and str(RT) not in str(existing_RT_value):
                                    new_ws.cell(row=1, column=x).value = str(existing_RT_value) + ', ' + str(RT)
                                    new_ws.cell(row=1, column=x).alignment = center_aligned_text
                        x = x + 1
                line_number += 1
        y += 1
    ### End of this section ###

    y = 5
    invalid_file_count = 0
    invalid_file_list = []
    for count in range(file_count):
        x = 3
        invalid_file_exist = False

        ##### v9 code #####
        for peak_index in range(peak_number):
            Name_ws = excel_name_list[peak_index]  # for comparison purpose
            RRT_ws = excel_RRT_list[peak_index]  # for comparison purpose
            ##### v9 code #####

            ##### v8 code #####
            # for peak_index in range(peak_number):
            #     Name_RRT_tuple = Name_RRT_list_final[peak_index]
            #     Name_ws = Name_RRT_tuple[0]  # for comparison purpose
            #     RRT_ws = Name_RRT_tuple[1]  # for comparison purpose
            ##### v8 code #####

            if RRT_ws == '1.00' or RRT_ws == '1.000' or RRT_ws == '1':
                if new_ws.cell(row=y, column=x).value == None and 'RRT~ 1.00' not in Name_ws:
                    invalid_file_exist = True
                    invalid_file_count += 1
                    new_ws.cell(row=y, column=x).value = 'No data'
                    new_ws.cell(row=y, column=x).font = Font(color="00FF0000")
                    new_ws.cell(row=y, column=x).alignment = center_aligned_text
                    # print(f'x-position: {x}, y-position: {y}')
            x += 1
        if invalid_file_exist == True:  # Highligh sample without data at RRT = 1
            # new_ws.cell(row=y, column=3).value = 'No AN value for RRT = 1'
            # new_ws.cell(row=y, column=3).font = Font(color="00FF0000", italic = True)
            new_ws.cell(row=y, column=2).font = Font(color="00FF0000")
            new_ws.cell(row=y, column=2).comment = Comment('RRT 1.00 has no AN value', 'Timothy Chang')
            new_ws.cell(row=y, column=2).comment.width = 200
            new_ws.cell(row=y, column=2).comment.height = 20
            new_ws.cell(row=y, column=1).font = Font(color="00FF0000")
            invalid_file_list.append(new_ws.cell(row=y, column=1).value)
        y += 1
    print(f'The invalid_file_list is: {invalid_file_list}\n')

    ## This section reorganize the presentation of headers after sorting the values in ascending order ##
    x_header = 3
    lam_list = []  # lambda list
    RT_list = []  # RT list
    RRT_list.clear()  # RRT list
    delimit = ','
    for peak_index in range(peak_number):
        lam_value = new_ws.cell(row=2, column=x_header).value
        RT_value = new_ws.cell(row=1, column=x_header).value
        RRT_value = new_ws.cell(row=3, column=x_header).value
        # print(f'RRT value one by one is: {RRT_value}')
        # print(f'x-position is: {x_header}.')
        Name_value = new_ws.cell(row=4, column=x_header).value
        if RRT_value == 1:
            new_ws.cell(row=3, column=x_header).font = Font(bold=True, color="000000FF")
        if 'GS-' in Name_value:
            new_ws.cell(row=4, column=x_header).font = Font(bold=True, color="000000FF")

        if type(lam_value) is str:
            sorted_lam_value = set(sorted(lam_value.split(', ')))
            sorted_lam_value = sorted(sorted_lam_value)
            for i in range(len(sorted_lam_value)):
                lam = float(sorted_lam_value[i])
                lam = "{:.1f}".format(lam)
                lam_list.append(lam)

            lam_list = set(lam_list)
            lam_list = sorted(lam_list)
            if len(lam_list) == 1:
                new_ws.cell(row=2, column=x_header).value = "{:.1f}".format(lam_list[0])
                new_ws.cell(row=2, column=x_header).alignment = center_aligned_text
            else:
                lam_list_str = str(lam_list)[1:-1]
                lam_list_str = lam_list_str.replace("'", "")
                lam_list_value = delimit.join(lam_list_str.split(','))
                new_ws.cell(row=2, column=x_header).value = lam_list_value
                new_ws.cell(row=2, column=x_header).alignment = center_aligned_text
            lam_list.clear()

        if type(RT_value) is str:
            sorted_RT_value = set(sorted(RT_value.split(', ')))
            sorted_RT_value = sorted(sorted_RT_value)
            for i in range(len(sorted_RT_value)):
                RT = float(sorted_RT_value[i])
                RT_list.append(RT)

            RT_list = set(RT_list)
            RT_list = sorted(RT_list)
            if len(RT_list) == 1:
                new_ws.cell(row=1, column=x_header).value = RT_list[0]
                new_ws.cell(row=1, column=x_header).alignment = center_aligned_text
            else:
                # RT_list_value = str(RT_list[0]) + ' - ' + str(RT_list[-1])
                RT_list_value = "{:.2f}".format(RT_list[0]) + ' - ' + "{:.2f}".format(RT_list[-1])
                new_ws.cell(row=1, column=x_header).value = RT_list_value
                new_ws.cell(row=1, column=x_header).alignment = center_aligned_text
                # new_ws.cell(row=1, column=x_header).number_format = numbers.FORMAT_NUMBER_00
                # RT_list_str = str(RT_list)[1:-1]  # purpose is for printing to Excel without brackets
                # RT_list_value = delimit.join(RT_list_str.split(','))
                # new_ws.cell(row=1, column=x_header).value = RT_list_value
                # new_ws.cell(row=1, column=x_header).alignment = center_aligned_text
            RT_list.clear()

        if type(RRT_value) is str:
            # print(Name_value)
            # print(RRT_value)
            # print()
            sorted_RRT_value = set(sorted(RRT_value.split(', ')))
            sorted_RRT_value = sorted(sorted_RRT_value)
            for i in range(len(sorted_RRT_value)):
                RRT = float(sorted_RRT_value[i])
                RRT = "{:.2f}".format(RRT)
                RRT_list.append(RRT)

            RRT_list = set(RRT_list)
            RRT_list = sorted(RRT_list)
            if len(RRT_list) == 1:
                new_ws.cell(row=3, column=x_header).value = RRT_list[0]
                new_ws.cell(row=3, column=x_header).alignment = center_aligned_text
            else:
                RRT_list_str = str(RRT_list)[1:-1]
                RRT_list_str = RRT_list_str.replace("'", "")
                RRT_list_value = delimit.join(RRT_list_str.split(','))
                new_ws.cell(row=3, column=x_header).value = RRT_list_value
                new_ws.cell(row=3, column=x_header).alignment = center_aligned_text
            RRT_list.clear()
        else:
            new_ws.cell(row=3, column=x_header).number_format = numbers.FORMAT_NUMBER_00

        if new_ws.cell(row=3, column=x_header).value == new_ws.cell(row=3,
                                                                    column=x_header - 1).value:  # Highlight peak with same RRT; help user tidy up data
            # new_ws.cell(row=3, column=x).font = Font(color="00FF6600")
            new_ws.cell(row=3, column=x_header).font = Font(bold=True, color="00FF6600")
            # new_ws.cell(row=3, column=x-1).font = Font(color="00FF6600")
            new_ws.cell(row=3, column=x_header - 1).font = Font(bold=True, color="00FF6600")
        x_header += 1
    ## End of this subsection ##

    # print(f'Data types are: {data_type}')

    # This section deals with sample without AN value at RRT = 1
    y = 5  # starting row for file name and sample name population in Excel
    x = 3  # starting column for AN population in Excel
    Unassigned_data = False
    file_count = len(file_sample_list_final)
    invalid_file_count = len(invalid_file_list)
    excel_RT_count = len(excel_RRT_list)
    for i in range(file_count):
        file_name = new_ws.cell(row=y, column=1).value
        if file_name in Do_Not_Overwrite_list:  # handles special condition where after QL is applied, AN value for RRT = 1 becomes ND
            with open(file_name) as csv_file:
                csv_reader = csv.reader(csv_file, delimiter='\t')
                line_number = 0
                for row in csv_reader:
                    if line_number > 2:
                        Name = row[0]
                        RRT = row[
                            1]  # handles special condition where after QL is applied, AN value for RRT = 1 becomes ND
                        AN = row[2]
                        x = 3
                        if RRT != '':
                            while x < excel_RT_count + 3:
                                Name_excel = new_ws.cell(row=4, column=x).value
                                if Name == Name_excel and float(RRT) == 1:
                                    message = 'Due to QL setting of ' + str(QL) + '; AN value was ' + str(AN) + '.'
                                    new_ws.cell(row=y, column=x).comment = Comment(message, 'Timothy Chang')
                                    new_ws.cell(row=y, column=x).comment.width = 300
                                    new_ws.cell(row=y, column=x).comment.height = 20
                                x += 1
                    line_number += 1
        elif file_name not in Do_Not_Overwrite_list:  # handles special condition where after QL is applied, AN value for RRT = 1 becomes ND
            for i_invalid in range(invalid_file_count):
                if file_name == invalid_file_list[i_invalid]:
                    with open(file_name) as csv_file:
                        csv_reader = csv.reader(csv_file, delimiter='\t')
                        line_number = 0  # First line of txt file.
                        for row in csv_reader:
                            if line_number == 2:  # Count number of column in txt file
                                column_count = len(row)
                                data_type = row
                                # if 'Peak Lambda Max.' in data_type:
                                #     lambda_max_exist = True
                                # else:
                                #     lambda_max_exist = False

                                # if 'Retention Time' in data_type:
                                #     RT_exist = True
                                # else:
                                #     RT_exist = False

                            if line_number > 2:
                                Name = row[0]
                                # RRT = row[1]  # By default, invalid file does not have RRT value because no AN value for RRT = 1; unless the special case where QL setting is the cause
                                AN = row[2]
                                if AN != '':
                                    if float(AN) < float(QL):
                                        AN = ''
                                        # print(AN)
                                lambda_max = row[3]
                                RT = row[4]
                                # if 'RRT~' in row[0]:
                                #    Name = 'RRT~ ' + row[1]

                                if AN != '' and RT != '':
                                    x = 3
                                    AN_assigned = False
                                    while x < excel_RT_count + 3 and AN_assigned == False:
                                        cell_occupied = False
                                        RT_excel = new_ws.cell(row=1, column=x).value
                                        Lambda_excel = new_ws.cell(row=2, column=x).value
                                        Name_excel = new_ws.cell(row=4, column=x).value
                                        # Assign AN value based on matching peak name
                                        if Name == Name_excel:
                                            AN_assigned = True
                                            if new_ws.cell(row=y,
                                                           column=x).value != None:  # Deal with more than one value occupying same cell and highlight in red font
                                                cell_occupied = True  # Cell already has a value (ie occupied)
                                            else:
                                                cell_occupied = False
                                        # Assign AN value based on RT
                                        elif type(RT_excel) == float and RT_excel == float(RT):
                                            if str(lambda_max) in str(
                                                    Lambda_excel):  # Cross check with lambda max
                                                AN_assigned = True
                                                if new_ws.cell(row=y,
                                                               column=x).value != None:  # Deal with more than one value occupying same cell and highlight in red font
                                                    cell_occupied = True  # Cell already has a value (ie occupied)
                                                else:
                                                    cell_occupied = False
                                                    # Assign AN value based on RT range
                                        elif type(RT_excel) == str:
                                            RT_range = RT_excel.split(' - ')
                                            RT_low_limit = float(RT_range[0])
                                            RT_high_limit = float(RT_range[1])
                                            RT = float(RT)
                                            # print(f'RT_low_limit is: {RT_low_limit}. RT_high_limit is: {RT_high_limit}')
                                            # print(type(RT))
                                            if RT >= RT_low_limit and RT <= RT_high_limit:
                                                if str(lambda_max) in str(Lambda_excel):  # Cross check with lambda max
                                                    AN_assigned = True
                                                    # print('true')
                                                    if new_ws.cell(row=y,
                                                                   column=x).value != None:  # Deal with more than one value occupying same cell and highlight in red font
                                                        cell_occupied = True  # Cell already has a value (ie occupied)
                                                        # print(row)
                                                        # print(f'Warning; cell already has a value, which is {new_ws.cell(row=y, column=x).value}. File name is: {file_name}\n')
                                                        # existing_value = new_ws.cell(row=y, column=x).value
                                                        # overlap_value = float(AN)
                                                        # new_ws.cell(row=y, column=x).value = str(existing_value) + ', ' + str(overlap_value)
                                                        # new_ws.cell(row=y, column=x).font = Font(color="00808080")
                                                        # new_ws.cell(row=y, column=x).number_format = numbers.FORMAT_NUMBER_00
                                                    else:
                                                        cell_occupied = False
                                                        # new_ws.cell(row=y, column=x).value = float(AN)
                                                        # new_ws.cell(row=y, column=x).number_format = numbers.FORMAT_NUMBER_00
                                                        # new_ws.cell(row=y, column=x).font = Font(color="00808080")
                                        else:
                                            AN_assigned = False

                                        # Print AN value to Excel
                                        if cell_occupied == True and AN_assigned == True:  # Deal with more than one value occupying the same cell
                                            # print(row)
                                            # print(f'Warning; cell already has a value, which is {new_ws.cell(row=y, column=x).value}. File name is: {file_name}\n')
                                            existing_value = new_ws.cell(row=y, column=x).value
                                            overlap_value = float(AN)
                                            new_ws.cell(row=y, column=x).value = str(existing_value) + ', ' + str(
                                                overlap_value)
                                            new_ws.cell(row=y, column=x).font = Font(color="00808080")
                                            new_ws.cell(row=y, column=x).number_format = numbers.FORMAT_NUMBER_00
                                        elif cell_occupied == False and AN_assigned == True:
                                            new_ws.cell(row=y, column=x).value = float(AN)
                                            new_ws.cell(row=y, column=x).number_format = numbers.FORMAT_NUMBER_00
                                            new_ws.cell(row=y, column=x).font = Font(color="00808080")
                                        x += 1

                                        # This section deals with AN value that does not fit into the table (No RT category in row 1 satisfied)
                                        if AN_assigned == False and x == excel_RT_count + 3:
                                            Unassigned_data = True
                                            if lambda_max != '':
                                                comment = 'RT ' + str(RT) + ': ' + str(AN) + ' (' + str(
                                                    lambda_max) + ' nm)'
                                            else:
                                                comment = 'RT ' + str(RT) + ': ' + str(AN)

                                            if new_ws.cell(row=y, column=x).value != None:
                                                existing_comment = new_ws.cell(row=y, column=x).value
                                                updated_comment = existing_comment + '; ' + comment
                                                new_ws.cell(row=y, column=x).value = updated_comment
                                                new_ws.cell(row=y, column=x).font = Font(color="00808080")
                                            else:
                                                new_ws.cell(row=y, column=x).value = comment
                                                new_ws.cell(row=y, column=x).font = Font(color="00808080")

                            line_number += 1
            y += 1
        else:  # handles special condition where after QL is applied, AN value for RRT = 1 becomes ND
            y += 1
            # End of this section

    y = file_count + 7
    new_ws.cell(row=y, column=1).value = 'data_compiler_v12'
    new_ws.cell(row=y, column=1).font = Font(bold=True)

    y += 1
    x = excel_RT_count + 3
    tabulated_file_count = file_count - invalid_file_count
    print(f'Number of file without a value for RRT = 1: {invalid_file_count}')
    new_ws.cell(row=y, column=1).value = '  [' + str(tabulated_file_count) + '/' + str(
        file_count) + ' files tabulated by RRT]'
    if invalid_file_count != 0:
        # new_ws.cell(row=y+1, column=1).value = '  [Please check data that cannot be assigned based on RT (row 1) or Lambda Max (row 2)]'
        new_ws.cell(row=y + 1,
                    column=1).value = '  [Check data for file highlighted in RED. AN value cannot be assigned based on RRT; no AN value for RRT = 1]'
        if Unassigned_data == True:
            new_ws.cell(row=y + 2,
                        column=1).value = '  [See Comment Column for AN value that cannot be assigned based on Peak Name (row 4), RT or RT range (row 1)]'
            new_ws.cell(row=4, column=x).value = 'Comment'
            new_ws.cell(row=4, column=x).font = Font(bold=True)
            new_ws.cell(row=4, column=x).alignment = center_aligned_text

    new_ws.cell(row=1, column=1).value = 'QL = ' + str(QL)
    new_ws.cell(row=1, column=1).font = Font(bold=True)
    new_ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    ##### Section for Formatting in a separate worksheet named "Formatted" #####
    ws2 = wb.create_sheet('Formatted')
    # y = 5 # starting row for file name and sample name population in Excel
    # x = 3 # starting column for AN population in Excel
    table_dimen_x = len(excel_name_list)
    table_dimen_y = len(file_sample_list_final)

    # This section finds column index where the name of peak begins with RRT~
    name_RRT_index = []
    GS_name_index = []
    x = 3  # starting position of RRT or Name header
    while x < table_dimen_x + 3:
        peak_name = new_ws.cell(row=4, column=x).value
        if 'RRT~' in str(peak_name):
            name_RRT_index.append(x)
        else:
            GS_name_index.append(x)
        x += 1
    ws2_dimen_x = len(GS_name_index)

    ws2.cell(row=1, column=1).value = 'QL = ' + str(QL)
    ws2.cell(row=1, column=1).font = Font(bold=True)
    ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    ws2.cell(row=4, column=ws2_dimen_x + 3).value = 'Unspecified Impurities'
    ws2.cell(row=4, column=ws2_dimen_x + 3).font = Font(bold=True)
    ws2.cell(row=4, column=ws2_dimen_x + 3).alignment = Alignment(horizontal='center')

    ws2.cell(row=4, column=ws2_dimen_x + 4).value = 'Unassigned AN Values'
    ws2.cell(row=4, column=ws2_dimen_x + 4).font = Font(bold=True)
    ws2.cell(row=4, column=ws2_dimen_x + 4).alignment = Alignment(horizontal='center')
    ws2.cell(row=4, column=ws2_dimen_x + 4).comment = Comment(
        'Unable to assign AN value based on RRT and Peak Name; no value for RRT = 1.  Assignment to the table was attempted by Data_Compiler if the peak satisfied both criteria of matching RT/RT range (row 1) AND Lambda Max (row 2).',
        'Timothy Chang')
    ws2.cell(row=4, column=ws2_dimen_x + 4).comment.width = 300
    ws2.cell(row=4, column=ws2_dimen_x + 4).comment.height = 100

    # This section print title headers (file name, sample name, RRT and Peak Name)
    ws2.cell(row=4, column=1).value = new_ws.cell(row=4, column=1).value
    ws2.cell(row=4, column=1).font = Font(bold=True)
    ws2.cell(row=4, column=1).alignment = Alignment(horizontal='left')

    ws2.cell(row=4, column=2).value = new_ws.cell(row=4, column=2).value
    ws2.cell(row=4, column=2).font = Font(bold=True)
    ws2.cell(row=4, column=2).alignment = right_aligned_text

    ws2.cell(row=1, column=2).value = new_ws.cell(row=1, column=2).value
    ws2.cell(row=1, column=2).font = Font(bold=True)
    ws2.cell(row=1, column=2).alignment = right_aligned_text

    ws2.cell(row=2, column=2).value = new_ws.cell(row=2, column=2).value
    ws2.cell(row=2, column=2).font = Font(bold=True)
    ws2.cell(row=2, column=2).alignment = right_aligned_text

    ws2.cell(row=3, column=2).value = new_ws.cell(row=3, column=2).value
    ws2.cell(row=3, column=2).font = Font(bold=True)
    ws2.cell(row=3, column=2).alignment = right_aligned_text

    ws2.cell(row=4, column=2).value = new_ws.cell(row=4, column=2).value
    ws2.cell(row=4, column=2).font = Font(bold=True)
    ws2.cell(row=4, column=2).alignment = right_aligned_text

    # This section prints headers (RT, Lambda Max, RRT and Name) to new worksheet 'Formatted'
    i = 0
    x_ws2 = 3
    while i < len(GS_name_index):
        x = GS_name_index[i]
        RT_header = new_ws.cell(row=1, column=x).value
        Lambda_header = new_ws.cell(row=2, column=x).value
        RRT_header = new_ws.cell(row=3, column=x).value
        Name_header = new_ws.cell(row=4, column=x).value
        ws2.cell(row=1, column=x_ws2).value = RT_header
        ws2.cell(row=2, column=x_ws2).value = Lambda_header
        ws2.cell(row=3, column=x_ws2).value = RRT_header
        ws2.cell(row=4, column=x_ws2).value = Name_header

        ws2.cell(row=1, column=x_ws2).alignment = center_aligned_text
        ws2.cell(row=2, column=x_ws2).alignment = center_aligned_text
        ws2.cell(row=3, column=x_ws2).font = Font(bold=True)
        ws2.cell(row=3, column=x_ws2).alignment = center_aligned_text
        ws2.cell(row=1, column=x_ws2).number_format = numbers.FORMAT_NUMBER_00
        ws2.cell(row=2, column=x_ws2).number_format = '0.0'
        ws2.cell(row=3, column=x_ws2).number_format = numbers.FORMAT_NUMBER_00
        ws2.cell(row=4, column=x_ws2).font = Font(bold=True)
        ws2.cell(row=4, column=x_ws2).alignment = center_aligned_text

        # This subsection prints AN values
        y = 5
        y_ws2 = 5
        while y < table_dimen_y + 5:
            message = new_ws.cell(row=y, column=x).comment
            ws2.cell(row=y_ws2, column=x_ws2).comment = message

            AN_value = new_ws.cell(row=y, column=x).value
            if AN_value == None:
                ws2.cell(row=y_ws2, column=x_ws2).value = 'ND'
                ws2.cell(row=y_ws2, column=x_ws2).alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws2.cell(row=y_ws2, column=x_ws2).value = AN_value
                ws2.cell(row=y_ws2, column=x_ws2).alignment = Alignment(horizontal='center', vertical='center')
                ws2.cell(row=y_ws2, column=x_ws2).number_format = numbers.FORMAT_NUMBER_00
            y_ws2 += 1
            y += 1
        x_ws2 += 1
        i += 1

    # This section print unassigned AN value to last column on the right
    y = 5
    y_ws2 = 5
    emDash = u'\u2014'
    while y < table_dimen_y + 5:
        unassigned_AN = new_ws.cell(row=y, column=table_dimen_x + 3).value
        if unassigned_AN == None:
            ws2.cell(row=y_ws2, column=ws2_dimen_x + 4).value = emDash
            ws2.cell(row=y_ws2, column=ws2_dimen_x + 4).alignment = Alignment(horizontal='center', vertical='center')
        else:
            ws2.cell(row=y_ws2, column=ws2_dimen_x + 4).value = unassigned_AN
            ws2.cell(row=y_ws2, column=ws2_dimen_x + 4).alignment = Alignment(horizontal='left', vertical='center')

        if len(name_RRT_index) == 0:  # This subsection deal with scenario where every peak has a none RRT name
            ws2.cell(row=y_ws2, column=ws2_dimen_x + 3).value = emDash
            ws2.cell(row=y_ws2, column=ws2_dimen_x + 3).alignment = Alignment(horizontal='center', vertical='center')
        y_ws2 += 1
        y += 1

    # Unspecified Impurities Section
    # This section print AN values in a single cell (wrapped text style) if name begins with RRT~
    y = 5
    while y < table_dimen_y + 5:
        sample_name = new_ws.cell(row=y, column=2).value  # Transfer sample name from "Compiled Data" worksheet
        file_name = new_ws.cell(row=y, column=1).value

        ws2.cell(row=y, column=2).value = sample_name  # Print sample name to "Formatted" worksheet
        ws2.cell(row=y, column=2).font = Font(bold=True)
        ws2.cell(row=y, column=2).alignment = Alignment(vertical='center')

        ws2.cell(row=y, column=1).value = file_name  # Print file name to "Formatted" worksheet
        ws2.cell(row=y, column=1).alignment = Alignment(vertical='center')

        font_color = new_ws.cell(row=y,
                                 column=2).font.color.rgb  # Change sample name and file name to red if no AN value for RRT = 1
        if font_color == '00FF0000':
            ws2.cell(row=y, column=1).font = Font(color=font_color)
            ws2.cell(row=y, column=2).font = Font(bold=True, color=font_color)
            ws2.cell(row=y, column=2).comment = Comment('RRT 1.00 has no AN value', 'Timothy Chang')
            ws2.cell(row=y, column=2).comment.width = 200
            ws2.cell(row=y, column=2).comment.height = 20

        i = 0
        while i < len(name_RRT_index):
            x = name_RRT_index[i]
            RRT_header = new_ws.cell(row=3, column=x).value
            RRT_header = format(RRT_header, '.2f')
            peak_value = new_ws.cell(row=y, column=x).value
            if peak_value != None:
                if ws2.cell(row=y, column=ws2_dimen_x + 3).value == None:
                    if type(peak_value) == str:  # handles special case where multiple values in one AN value cell in main table
                        ws2.cell(row=y, column=ws2_dimen_x + 3).value = 'RRT ' + str(RRT_header) + ': ' + peak_value
                        ws2.cell(row=y, column=ws2_dimen_x + 3).alignment = center_aligned_text
                        ws2.cell(row=y, column=ws2_dimen_x + 3).alignment = Alignment(wrapText=True,
                                                                                      horizontal='center')
                    else:
                        peak_value = format(peak_value, '.2f')
                        ws2.cell(row=y, column=ws2_dimen_x + 3).value = 'RRT ' + str(RRT_header) + ': ' + str(
                            peak_value)
                        ws2.cell(row=y, column=ws2_dimen_x + 3).alignment = center_aligned_text
                else:  # handles the scenario where an RRT entry is already in the Unspecified Impurities cell
                    if type(peak_value) == str:  # handles special case where multiple values in one AN value cell in main table
                        # print('executed')
                        ws2.cell(row=y, column=ws2_dimen_x + 3).value = 'RRT ' + str(RRT_header) + ': ' + peak_value
                        ws2.cell(row=y, column=ws2_dimen_x + 3).alignment = center_aligned_text
                        ws2.cell(row=y, column=ws2_dimen_x + 3).alignment = Alignment(wrapText=True,
                                                                                      horizontal='center')
                    else:
                        prior_cell_content = ws2.cell(row=y, column=ws2_dimen_x + 3).value
                        peak_value = format(peak_value, '.2f')
                        new_cell_content = prior_cell_content + '\n' + 'RRT ' + str(RRT_header) + ': ' + str(peak_value)
                        ws2.cell(row=y, column=ws2_dimen_x + 3).value = new_cell_content
                        ws2.cell(row=y, column=ws2_dimen_x + 3).alignment = center_aligned_text
                        ws2.cell(row=y, column=ws2_dimen_x + 3).alignment = Alignment(wrapText=True,
                                                                                      horizontal='center')
            i += 1
            # This subsection insert an emDash if Unspecified column is empty
            if i == len(name_RRT_index) and ws2.cell(row=y, column=ws2_dimen_x + 3).value == None:
                emDash = u'\u2014'
                ws2.cell(row=y, column=ws2_dimen_x + 3).value = emDash
                ws2.cell(row=y, column=ws2_dimen_x + 3).alignment = center_aligned_text
        y += 1

    y = table_dimen_y + 7
    ws2.cell(row=y, column=1).value = 'data_compiler_v12'
    ws2.cell(row=y, column=1).font = Font(bold=True)

    wb.save('Tabulated_v12.xlsx')





